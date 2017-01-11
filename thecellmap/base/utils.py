'''
Created on Dec 13, 2013

@author: matej
'''
from StringIO import StringIO
import csv
import datetime
import fileinput
import hotshot
import json
import os
import re
import tempfile
import time
from types import NoneType

from django.core.files.temp import NamedTemporaryFile
from django.core.management.base import BaseCommand, CommandError
from django.db import transaction
from django.http.response import HttpResponse
import openpyxl
import xlrd
from xlrd.biffh import XLRDError
from xlwt import Style
import xlwt

from base.models import Gene
import numpy as np
from thecellmap import settings


RE_ORF = re.compile("^Y[A-P][LR]\d{3}[CW](\-[A-Z])?$")

xlwt.add_palette_colour("red_stringent", 0x21)
xlwt.add_palette_colour("red_lenient", 0x22)
xlwt.add_palette_colour("green_stringent", 0x23)
xlwt.add_palette_colour("green_lenient", 0x24)
xlwt.add_palette_colour("correlation", 0x25)

STYLE_NEG_STRINGENT = '-str'
STYLE_NEG_SIGNIFICANT = '-sig'
STYLE_POS_STRINGENT = '+str'
STYLE_POS_SIGNIFICANT = '+sig'
STYLE_COR_SIGNIFICANT = 'cor'
STYLE_BOLD = 'bold'
STYLE_WRAP = 'wrap'

STYLES = {
    STYLE_NEG_STRINGENT: ('red_stringent', 'FF0000BF'),
    STYLE_NEG_SIGNIFICANT: ('red_lenient', 'FF0000FF'),
    STYLE_POS_STRINGENT: ('green_stringent', 'FFBFBF00'),
    STYLE_POS_SIGNIFICANT: ('green_lenient', 'FFFFFF00'),
    STYLE_COR_SIGNIFICANT: ('correlation', 'FFFFCC00'),
}

def get_xlwt_style(color):
    style = xlwt.XFStyle()
    pattern = xlwt.Pattern()
    pattern.pattern = xlwt.Pattern.SOLID_PATTERN
    pattern.pattern_fore_colour = xlwt.Style.colour_map[color]
    style.pattern = pattern
    return style

for typ in STYLES.keys():
    xc, xxc = STYLES[typ]
    STYLES[typ] = (get_xlwt_style(xc), xxc)

STYLES[STYLE_BOLD] = (xlwt.easyxf("font: bold on;"), STYLE_BOLD)
STYLES[STYLE_WRAP] = (xlwt.easyxf("align: wrap on, vert centre;"), STYLE_WRAP)

INSTRUCTIONS = 'instructions'

class CellMapCommand(BaseCommand):
    def get_path(self, filepath, required=True):
        if not required and filepath is None:
            return None
        if filepath is None:
            raise CommandError('File argument is required')
        
        if '~' in filepath:
            filepath = os.path.expanduser(filepath)
        
        return os.path.abspath(filepath)
        
    def get_fd(self, filepath, mode='r', required=True):
        filepath = self.get_path(filepath, required)
        if not filepath: return None
        
        if mode == 'r' and not os.path.isfile(filepath):
            raise CommandError('%s is not a file' % filepath)
        
        if filepath.endswith('.gz'):
            return fileinput.hook_compressed(filepath, mode)
        return open(filepath, mode)

def orf_comparator(x,y):
    if not cmp(x[:2],y[:2]) and x[2] == 'L' and y[2] == 'L':
        return -cmp(x, y)
    return cmp(x, y)

def orf_sorting_value(orf):
    orf = orf.split('_', 1)[0]
    return ord(orf[1])*10000 + (1000 + (orf[2] == 'L' and -int(orf[3:6]) or int(orf[3:6])+1))

def ordered_orfs(orfs):
    return sorted(orfs, cmp=orf_comparator)

def gene_map(keyfun=lambda x: x):
    genemap = {}
    for g in Gene.objects.all():
        genemap[keyfun(g.orf)] = g
        if g.name:
            genemap[keyfun(g.name)] = g
        for a in g.aliases or []:
            if a and a not in genemap:
                genemap[keyfun(a)] = g
    return genemap

COLORS = {'blue':34, 'cyan':36, 'green':32, 'grey':30, 'magenta':35, 'red':31, 'white':37, 'yellow':33}
RESET = '\033[0m'

def print_queries(func, filter=None):
    """ Print all queries executed in this funnction. """
    def wrapper1(init_self, *args, **kwargs):
        from django.db import connection
        sqltime, longest, numshown = 0.0, 0.0, 0
        initqueries = len(connection.queries)
        starttime = time.time()
        result = func(init_self, *args, **kwargs)
        for query in connection.queries[initqueries:]:
            sqltime += float(query['time'].strip('[]s'))
            longest = max(longest, float(query['time'].strip('[]s')))
            if not filter or filter in query['sql']:
                numshown += 1
                querystr = colored('\n[%ss] ' % query['time'], 'yellow')
                querystr += colored(query['sql'], 'blue')
                print querystr
        numqueries = len(connection.queries) - initqueries
        numhidden = numqueries - numshown
        runtime = round(time.time() - starttime, 3)
        proctime = round(runtime - sqltime, 3)
        print colored("------", 'blue')
        print colored('Total Time:  %ss' % runtime, 'yellow')
        print colored('Proc Time:   %ss' % proctime, 'yellow')
        print colored('Query Time:  %ss (longest: %ss)' % (sqltime, longest), 'yellow')
        print colored('Num Queries: %s (%s hidden)\n' % (numqueries, numhidden), 'yellow')
        return result
    
    """ Use this only in debug mode (aka non production mode) """
    if settings.DEBUG:
        return wrapper1
    return func

def colored(text, color=None):
    """ Colorize text {red, green, yellow, blue, magenta, cyan, white}. """
    if os.getenv('ANSI_COLORS_DISABLED') is None:
        fmt_str = '\033[%dm%s'
        if color is not None:
            text = fmt_str % (COLORS[color], text)
        text += RESET
    return text

class GenericXlsWriter():
    def __init__(self, fd):
        self.fd = fd
        self.workbook = self._create_wb()
        self.sheets = {}
        self.sheet_ord = []
        self.active_sheet = None
    
    def _format_sheet_name(self, name):
        return name and name.replace(':', '-')[:31] or None
    
    def set_cur_sheet(self, sheet):
        self.active_sheet = sheet
    def get_cur_sheet(self):
        return self.active_sheet
    cur_sheet = property(get_cur_sheet, set_cur_sheet)
    
    def reset_row(self, row=0, sheet=None):
        self._write_get_sheet(sheet)['row'] = row
    
    def add_sheet(self, name, headers=None):
        name = self._format_sheet_name(name)
        
        if name in self.sheets:
            raise Exception('Sheet %s exists' % name)
        
        # TODO: name length fix and give warning as return value
        self.sheets[name] = {'sheet': self._add_sheet(name), 'row': 0}
        self.sheet_ord.append(name)
        self.active_sheet = name
        
        if headers:
            self.write_row(headers, sheet=name, style=STYLE_BOLD)
    
    def write(self, row, col, value, sheet=None, **kwargs):
        sheet = self._format_sheet_name(sheet)
        
        if isinstance(sheet, (int, )):
            sheet = self.sheets.keys()[sheet]
        elif isinstance(sheet, (NoneType, )):
            sheet = self.active_sheet
        
        if sheet not in self.sheets:
            raise Exception('Sheet %s not in existing sheets' % sheet)
        
        sheet = self.sheets[sheet]
        self._write_cell(sheet['sheet'], row, col, value, **kwargs)
    
    def _write_get_sheet(self, sheet):
        sheet = self._format_sheet_name(sheet)
        
        if isinstance(sheet, (int, )):
            sheet = self.sheets.keys()[sheet]
        elif isinstance(sheet, (NoneType, )):
            sheet = self.active_sheet
        
        if sheet not in self.sheets:
            raise Exception('Sheet %s not in existing sheets' % sheet)
        
        return self.sheets[sheet]
    
    def write_row(self, values, sheet=None, **kwargs):
        sheet = self._write_get_sheet(sheet)
        for col, val in enumerate(values):
            self._write_cell(sheet['sheet'], sheet['row'], col, val, **kwargs)
        sheet['row'] += 1
    
    writerow = write_row
    
    CORRELATION_FORMATS = (None, None, '0.000')
    def write_correlation_row(self, values, sheet=None, **kwargs):
        sheet = self._write_get_sheet(sheet)
        for col, val in enumerate(values):
            self._write_cell(sheet['sheet'], sheet['row'], col, val, number_format=self.CORRELATION_FORMATS[col], **kwargs)
        sheet['row'] += 1
    
    SCORE_FORMATS = (None, None, '0.000', '0.00E+00')
    def write_score_row(self, values, sheet=None, col_offset=0, **kwargs):
        sheet = self._write_get_sheet(sheet)
        for col, val in enumerate(values):
            self._write_cell(sheet['sheet'], sheet['row'], col + col_offset, val, number_format=self.SCORE_FORMATS[col], **kwargs)
        sheet['row'] += 1
    
    def write_score_row_neg(self, values, sheet=None, **kwargs):
        self.write_score_row(values, sheet, col_offset=0, **kwargs)
    
    def write_score_row_pos(self, values, sheet=None, **kwargs):
        self.write_score_row(values, sheet, col_offset=len(self.SCORE_FORMATS) + 1, **kwargs)
    
    def add_instructions_sheet(self):
        self.add_sheet('Instructions')
    
    def write_instructions(self, content):
        sheet = self._write_get_sheet('Instructions')
        self._write_cell(sheet['sheet'], 0, 0, 'Source', style=STYLE_BOLD)
        self._write_cell(sheet['sheet'], 0, 1, 'thecellmap.org')
        self._write_cell(sheet['sheet'], 1, 0, 'Downloaded on', style=STYLE_BOLD)
        self._write_cell(sheet['sheet'], 1, 1, datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S'))
        self._write_cell(sheet['sheet'], 2, 0, 'Content', style=STYLE_BOLD)
        self._write_cell(sheet['sheet'], 2, 1, 'Genetic interactions scores (GI Scores) and similarity of genetic interaction profile similarity (Profile Sim.) for genes %s' % (content, ))
        
        self._write_cell(sheet['sheet'], 3, 0, 'Description', style=STYLE_BOLD)
        self._write_cell(sheet['sheet'], 4, 1, 'The SGA score measures the extent to which a double mutant colony size deviates from the colony size expected from combining two mutations together. '
'The SGA score measures the extent to which a double mutant colony size deviates from the colony size expected from combining two mutations together. '
'The tab labeled "GI Scores" includes negative (putative synthetic sick/lethal) and positive interactions (potential epistatic or suppression interactions) involving the gene(s) of interest. ' 
'The magnitude of the SGA score is indicative of the strength of interaction. '
'Based on statistical analysis, we determined default cutoffs for the quantitative genetic interactions. ' 
'The intermediate cutoff consists of a combination of p-value<0.05 and SGA score >|0.08|. ' 
'A more stringent threshold on negative interactions (p-value<0.05 and SGA score <-0.12) and positive interactions (p-value<0.05 and SGA score>0.16) are also indicated. '
'Note that none of these interactions are confirmed and are likely to include some false positives. ' 
'Thus, additional tests (e.g. random spore or tetrad analysis for negative interactions) should be performed prior to follow-up experiments. '
'Genes that share similar patterns of genetic interactions often belong to the same protein complex or biological pathway. '
'Thus, comparison of genetic interaction patterns/profiles is a powerful way to define gene function. '
'The attached spreadsheet includes a second tab "GI Profile Sim." which consists of a ranked list of genes whose genetic interaction patterns most closely resemble the gene of interest. ' 
'Genetic profile similarity is based on Pearson correlation.', style=STYLE_WRAP)

#         self._write_cell(sheet['sheet'], 3, 1, 'GI profile sim.', style=STYLE_BOLD)
#         self._write_cell(sheet['sheet'], 3, 2, 'Provides a ranked list of genes whose genetic interaction profile most closely resembles the genetic interaction profile of the gene(s) of interest.')
#         self._write_cell(sheet['sheet'], 4, 1, 'GI scores', style=STYLE_BOLD)
#         self._write_cell(sheet['sheet'], 4, 2, 'Lists the direct negative and positive interactions for the gene(s) of interest.')
#         
        self._write_cell(sheet['sheet'], 6, 0, 'Legend:', style=STYLE_BOLD)
        
        self._write_cell(sheet['sheet'], 7, 0, 'A', style=STYLE_COR_SIGNIFICANT)
        self._write_cell(sheet['sheet'], 7, 1, 'Genetic Interaction Profiles similar to gene of interest (Pearson Correlation Coefficient > 0.2)')
        self._write_cell(sheet['sheet'], 8, 0, 'B', style=STYLE_NEG_STRINGENT)
        self._write_cell(sheet['sheet'], 8, 1, 'Significant negative genetic interactions (stringent cutoff: score < -0.12, p-value < 0.05)')
        self._write_cell(sheet['sheet'], 9, 0, 'C', style=STYLE_NEG_SIGNIFICANT)
        self._write_cell(sheet['sheet'], 9, 1, 'Significant negative genetic interactions (intermediate cutoff: score < -0.08, p-value < 0.05)')
        self._write_cell(sheet['sheet'], 10, 0, 'D', style=STYLE_POS_STRINGENT)
        self._write_cell(sheet['sheet'], 10, 1, 'Significant positive genetic interactions (stringent cutoff: score > 0.16, p-value < 0.05)')
        self._write_cell(sheet['sheet'], 11, 0, 'E', style=STYLE_POS_SIGNIFICANT)
        self._write_cell(sheet['sheet'], 11, 1, 'Significant positive genetic interactions (intermediate cutoff: score > 0.08, p-value < 0.05)')
        
        sheet['sheet'].col(1).width = 24000
        sheet['sheet'].row(4).height = 4000
        
#         self._write_cell(sheet['sheet'], 11, 0, 'Notes:', style=STYLE_BOLD)
#         self._write_cell(sheet['sheet'], 12, 0, 'These are unpublished data. Please contact Michael Costanzo (michael.costanzo@utoronto.ca) for questions regarding citation policy.')
    
    def save(self, seek=None):
        self._save()
        if isinstance(seek, (int, )):
            self.fd.seek(seek)
        return self.fd
    
    def as_response(self):
        response = HttpResponse(content_type=self.mime)
        response['Content-Disposition'] = 'attachment; filename=%s' % (self.fd, )
        self.fd = response
        self.save()
        return response
    
    def sheets(self):
        return self.sheet_ord

class XlsWriter(GenericXlsWriter):
    mime = 'application/vnd.ms-excel'
    
    def _create_wb(self):
        wb = xlwt.Workbook()
        wb.set_colour_RGB(0x21, 0, 101, 204)
        wb.set_colour_RGB(0x22, 0, 118, 239)
        wb.set_colour_RGB(0x23, 205, 164, 0)
        wb.set_colour_RGB(0x24, 255, 204, 0)
        wb.set_colour_RGB(0x25, 255, 204, 153)
        return wb
    
    def _add_sheet(self, name):
        ws = self.workbook.add_sheet(name) 
        ws.paper_size_code = 1 # US Letter
        return ws
    
    def _write_cell(self, sheet, row, col, value, style=None, number_format=None):
        style = style and STYLES[style][0] or Style.default_style
        if number_format:
            style.num_format_str = number_format
        
        sheet.write(row, col, value, style)
    
    def _save(self):
        self.workbook.save(self.fd)

class XlsxWriter(GenericXlsWriter):
    mime = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    
    def _create_wb(self):
        wb = openpyxl.workbook.Workbook()
        wb.remove_sheet(wb.get_active_sheet())
        return wb
    
    def _add_sheet(self, name):
        ws = self.workbook.create_sheet()
        ws.set_printer_settings(ws.PAPERSIZE_LETTER, ws.ORIENTATION_PORTRAIT) # leave this here just in case... 
        ws.page_setup.paperSize = ws.PAPERSIZE_LETTER # this actually sets the size
        ws.title = name
        return ws
    
    def _write_cell(self, sheet, row, col, value, style=None, number_format=None):
        cell = sheet.cell(row=row+1, column=col+1)
        cell.value = value
        
#         if style == STYLE_BOLD:
#             cell.style.font.bold = True
#         elif style:
#             cell.style.fill.fill_type = Fill.FILL_SOLID
#             cell.style.fill.start_color.index = STYLES[style][1]
    
    def _save(self):
        self.workbook.save(self.fd)

class CsvWriter(GenericXlsWriter):
    mime = 'text/csv'
    
    def __init__(self, fd, delimiter):
        self.file_name = ''
        self.delimiter = delimiter
        GenericXlsWriter.__init__(self, fd)
        if delimiter == '\t':
            self.mime = 'text/tab-separated-values'
    
    def _create_wb(self):
        if not hasattr(self.fd, 'write'):
            self.file_name = self.fd
            self.fd = StringIO()
            
        return csv.writer(self.fd, delimiter=self.delimiter,
                                    quotechar='|', quoting=csv.QUOTE_MINIMAL)
    
    def _add_sheet(self, name):
        return self.workbook
    
    def write_row(self, values, sheet=None, **kwarg):
        sheet = self._write_get_sheet(sheet)
        sheet['sheet'].writerow(values)
    
    def as_response(self):
        if not self.file_name:
            raise Exception()
        
        response = HttpResponse(content_type=self.mime)
        response['Content-Disposition'] = 'attachment; filename=%s' % (self.file_name, )
        self.fd.seek(0)
        response.write(self.fd.read())
        return response

def write_excel_file(fd=None, type='xls', override_ext=False):
    if isinstance(fd, (str, unicode)) and override_ext:
        type = os.path.splitext(fd)[1].strip('.')
    elif fd == None:
        fd = StringIO()
    
    type = type.lower()
    
    if type == 'xls':
        return XlsWriter(fd)
    elif type == 'xlsx':
        return XlsxWriter(fd)
    elif type == 'csv':
        return CsvWriter(fd, ',')
    elif type == 'tsv':
        return CsvWriter(fd, '\t')
    
    raise BadXlsFile()

class GenericXls():
    def __init__(self, filename, sheet, name=None):
        self.name = name
        self.filename = filename
        self.workbook = self._open_workbook(filename)
        self.open_sheet(sheet)
    
    def _open_workbook(self, filename): raise NotImplementedError()
    def _open_sheet(self, name): raise NotImplementedError()
    def _get_row(self, row): raise NotImplementedError()
    def _get_value(self, cell): raise NotImplementedError()
    
    def open_sheet(self, sheet):
        if isinstance(sheet, int): sheet = self.sheets()[sheet]
        self.sheet = self._open_sheet(sheet)
        self.row = 0
    
    def skiplines(self, skip=1):
        self.row += skip
    
    def reset(self):
        self.row = 0
    
    def readline(self, min_columns=0, exact_columns=None):
        if self.row >= self.rows(): return None
        l = [self._get_value(c) for c in self._get_row(self.row)]
        while len(l) < min_columns:
            l.append(None)
        
        if exact_columns and len(l) != exact_columns:
            raise XlsError("Error on line %s: Wrong number of columns, expected %s, got %s instead" % (self.row, exact_columns, len(l)))
        
        self.row += 1
        return l
    
    def readlines(self, start=None, min_columns=None, exact_columns=None):
        if start:
            self.row = start
        
        while self.row < self.rows():
            yield self.readline(min_columns, exact_columns)
    
    def __iter__(self):
        for l in self.readlines():
            yield l
    
    def __getitem__(self, index):
        if isinstance(index, slice):
            self.row = index.start
            
            acc = []
            for l in self.readlines():
                if self.row >= index.stop:
                    break
                acc.append(l)
            return acc
        else:
            self.row = index
            return self.readline()
    
class Xls(GenericXls):
    def _open_workbook(self, filename):
        return xlrd.open_workbook(filename=filename)
    
    def _open_sheet(self, name):
        return self.workbook.sheet_by_name(name)
    
    def _get_row(self, row):
        return self.sheet.row(row)
    
    def rows(self):
        return self.sheet.nrows
    
    def sheets(self):
        return [s.name for s in self.workbook.sheets() if s.name.lower() != INSTRUCTIONS]
    
    def _get_value(self, cell):
        if cell.ctype == 3: # date
            return datetime.datetime(*xlrd.xldate_as_tuple(cell.value, self.workbook.datemode)).strftime('%Y-%m-%d')
        if cell.ctype == 5: # error
            return np.nan
        return cell.value
    
class Xlsx(GenericXls):
    def _open_workbook(self, filename):
        return openpyxl.reader.excel.load_workbook(filename)
    
    def _open_sheet(self, name):
        return self.workbook.get_sheet_by_name(name)
    
    def _get_row(self, row):
        return self.sheet.rows[row]

    def rows(self):
        return self.sheet.get_highest_row()
    
    def sheets(self):
        return [s for s in self.workbook.get_sheet_names() if s.lower() != INSTRUCTIONS]
    
    def _get_value(self, cell):
        return cell.value

class Csv(GenericXls):
    def _open_workbook(self, filename):
        with open(filename, 'rbU') as f:
            sniffer = csv.Sniffer()
            self.dialect = sniffer.sniff(f.readline())
            f.seek(0)
            self.numlines = len(f.readlines())
        
        return csv.reader(open(filename, 'rbU'), self.dialect)
    
    def _open_sheet(self, name): pass
    def _get_row(self, row): return self.workbook.next()
    def rows(self): return self.numlines
    def sheets(self): return ['csv']
    def _get_value(self, cell): return cell

def open_excel_file(filename, sheet=0, fd=None):
    if fd and not isinstance(fd, (str, unicode)):
        # This is stupid... oh well
        tmp = NamedTemporaryFile(delete=True)
        tmp.write(fd.read())
        tmp.seek(0)
        fd = tmp.name
    
    if filename.lower().endswith('xlsx'):
        return Xlsx(fd or filename, sheet=sheet, name=filename)
    elif filename.lower().endswith('xls'):
        try:
            return Xls(fd or filename, sheet=sheet, name=filename)
        except XLRDError:
            pass
    elif filename.lower().endswith('csv'):
        return Csv(fd or filename, sheet=sheet, name=filename)
    
    raise BadXlsFile()

class BadXlsFile(Exception): pass
class XlsError(Exception): pass

is_integer = lambda x: not not re.match('\d+', x)

class JsonResponse(HttpResponse):
    def __init__(self, obj):
        super(JsonResponse, self).__init__(json.dumps(obj), content_type="application/json")

try:
    PROFILE_LOG_BASE = settings.PROFILE_LOG_BASE  # @UndefinedVariable
except:
    PROFILE_LOG_BASE = tempfile.gettempdir()

def profile(log_file):
    """Profile some callable.

    This decorator uses the hotshot profiler to profile some callable (like
    a view function or method) and dumps the profile data somewhere sensible
    for later processing and examination.

    It takes one argument, the profile log name. If it's a relative path, it
    places it under the PROFILE_LOG_BASE. It also inserts a time stamp into the 
    file name, such that 'my_view.prof' become 'my_view-20100211T170321.prof', 
    where the time stamp is in UTC. This makes it easy to run and compare 
    multiple trials.     
    """

    if not os.path.isabs(log_file):
        log_file = os.path.join(PROFILE_LOG_BASE, log_file)

    def _outer(f):
        def _inner(*args, **kwargs):
            # Add a timestamp to the profile output when the callable
            # is actually called.
            (base, ext) = os.path.splitext(log_file)
            base = base + "-" + time.strftime("%Y%m%dT%H%M%S", time.gmtime())
            final_log_file = base + ext

            prof = hotshot.Profile(final_log_file)
            try:
                ret = prof.runcall(f, *args, **kwargs)
            finally:
                prof.close()
            return ret

        return _inner
    return _outer

def dump_clean_json(obj, f):
    with open(f, 'wb') as out:
        out.write(json.dumps(obj).replace(' ', ''))

def rollback_on_fail(fun):
   def wrap(init_self,*args,**kwargs):
       try:
           return fun(init_self,*args,**kwargs)
       except Exception:
           transaction.rollback()
           raise
   
   return wrap